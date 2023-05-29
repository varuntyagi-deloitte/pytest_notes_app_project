import openpyxl


def read_data_by_value(sheet_name, key):
    file = "C://Users//varutyagi//PycharmProjects//python_notes_app//TestData//Excel_file.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    row = sheet.max_row
    for i in range(1, row + 1):
        cell_value = sheet.cell(row=i, column=1).value
        if cell_value == key:
            return sheet.cell(row=i, column=2).value

